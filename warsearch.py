try:
    from googlesearch import search
except ImportError:
    print("No module named 'google' found")
try:
    from urllib.parse import urlparse
    import urllib.request
except ImportError:
    print("No module named 'urllib' found")
try:
    from pathlib import Path 
except ImportError:
    print("No module named 'pathlib' found")
try:
    from PyPDF2 import PdfFileReader 
except ImportError:
    print("No module named 'PyPDF2' found")
try:
    import os 
except ImportError:
    print("No module named 'os' found")
try:
    from datetime import datetime 
except ImportError:
    print("No module named 'datetime' found")
try:
    import openpyxl
except ImportError:
    print("No module named 'openpyxl' found")
try:
    import operator
except ImportError:
    print("No module named 'operator' found")

#classes
class AosUnit:
    def __init__(self, grand_alliance,  faction,  sub_faction,  unit):
        self.grand_alliance = grand_alliance
        self.faction = faction
        self.sub_faction = sub_faction
        self.unit = unit
        
class WarscrollURL:
    def __init__(self, host, url, path, name, ext, date,  check):
        self.url = url
        self.host = host
        self.name = name
        self.path = path
        self.ext = ext
        self.date = date
        self.check = check
    
# lists
gw_hostnames = [
    "whc-cdn.games-workshop.com",
    "www.forgeworld.co.uk", 
    "www.games-workshop.com",
    "www.warhammer-community.com"
    ]

# variables
import_list = []
good_urls = []
dl_path = "./downloads/"

# opener for downloading with urllib
opener = urllib.request.build_opener()
opener.addheaders = [('User-agent', 'Mozilla/5.0')]
urllib.request.install_opener(opener)

# import list of Warscrolls from XLSX
imp_wb = openpyxl.load_workbook("AoS Warscrolls short.xlsx")
imp_ws = imp_wb['Sheet1']
for r in range(2, imp_ws.max_row + 1):
    import_list.append(AosUnit(   imp_ws.cell(r, 1).value, 
                                                imp_ws.cell(r, 2).value, 
                                                imp_ws.cell(r, 3).value, 
                                                imp_ws.cell(r, 4).value
                                            )
                                )
imp_wb.close()

# create check list XLSX
# TODO - delete old checklist file
cl_wb = openpyxl.Workbook()
cl_wb.save('Warscrolls_to_Check.xlsx')
cl_ws = cl_wb.active
cl_ws.append(['Grand Alliance',  'Alliance',  'Sub Faction',  'Unit',  'Link',  'File Path', 'PDF Date'])

# create output XLSX (list of downloaded and sorted files)
# TODO - delete old output file
out_wb = openpyxl.Workbook()
out_wb.save('AoS Warscroll Database.xlsx')
out_ws = out_wb.active
out_ws.append(['Grand Alliance',  'Alliance',  'Sub Faction',  'Unit',  'Link',  'File Path', 'PDF Date'])

# Iterate through list of Warscrolls and create list of PDFs
#query = "Horrors of Tzeentch Warscroll"
for unit in import_list:
    query = unit.unit + " Warscroll"
    
    for url in search(query, tld="com", num=10, stop=10, pause=2):
        # parse result URL into components
        parsed = urlparse(url)
        
        # pick out intersting components
        host = parsed.hostname
        name = Path(parsed.path).name
        path = Path(parsed.path).parent
        ext = (Path(parsed.path).suffix).lower()
        
        # TODO - replace urllib with requests
        # if it's from Games Workshop & ends in 'pdf' save the url & download the file
        if host in gw_hostnames and ext == ".pdf":
            good_urls.append(WarscrollURL(host, url,  path, name,  ext, 0,  False))
            try:
                urllib.request.urlretrieve(url, os.path.join(dl_path, name))
            except (urllib.error.URLError,  urllib.error.HTTPError) as e:
                print("Downloader Reported an Error trying to download \"" + url + "\": " + str(e))
                

    # get the PDF creation date and update good_urls
    j = 1
    for i in good_urls:
        #TODO if file doesn't exist it is b/c the download couldn't download it (403 Forbidden for instance)... might need to fix this smarter
        if os.path.exists(os.path.join("./downloads/", i.name)):
            with open(os.path.join("./downloads/", i.name),  'rb') as f:
                pdf = PdfFileReader(f)
                info = pdf.getDocumentInfo()
                
                if(info['/CreationDate'][-1] == "Z"):
                    i.date = datetime.strptime(info['/CreationDate'], "D:%Y%m%d%H%M%SZ")
                else:
                    i.date = datetime.strptime(info['/CreationDate'], "D:%Y%m%d%H%M%S+01'00'")

        # rename files from warhammer-community & add note that they need to be checked
        if (i.host == "www.warhammer-community.com"):
            tempname = i.name
            i.name = query + " " + str(j) + i.ext
            os.remove(dl_path + i.name)
            os.rename(dl_path + tempname,  dl_path + i.name)
            i.check = True
            j = j + 1
    
    # create directory if necessary for newest file
    if not os.path.exists("./pdfs/" + unit.grand_alliance + "/" + unit.faction + "/" + unit.sub_faction):
        os.makedirs("./pdfs/" + unit.grand_alliance + "/" + unit.faction + "/" + unit.sub_faction)
    
    # create legacy directory if necessary
    if len(good_urls) > 1:
        if not os.path.exists("./pdfs/" + unit.grand_alliance + "/" + unit.faction + "/legacy"):
            os.makedirs("./pdfs/" + unit.grand_alliance + "/" + unit.faction + "/legacy")
    
    # sort good_urls by creation date
    good_urls.sort(key=operator.attrgetter('date'))
    
    # move newest file to proper location, move remaining files to legacy
    for i in good_urls:
        
        # move file from downloads folder to appropriate sub-folder in pdfs folder.
        if i == 0:     # newest file
            os.replace("./downloads/" + i.name, "./pdfs/" + unit.grand_alliance + "/" + unit.faction + "/" + unit.sub_faction + "/" + i.name)
        else:           # older files
            os.replace("./downloads/" + i.name,  "./pdfs/" + unit.grand_alliance + "/" + unit.faction + "/legacy/" + i.name)
        
        # append any warhammer-community files to the check list
        if (i.check):
            cl_ws.append([unit.grand_alliance,unit.faction, unit.sub_faction, i.url, str(i.path), i.date.strftime("%Y-%m-%d %H:%M:%S")])
            
        # append files to output XLSX
        cl_ws.append([unit.grand_alliance,unit.faction, unit.sub_faction, i.url, str(i.path), i.date.strftime("%Y-%m-%d %H:%M:%S")])

#close workbooks
cl_wb.close()
out_wb.close()
    
